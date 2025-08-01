from django.shortcuts import get_object_or_404, render

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from app_account.models import IPRegectedUser, SessionWarningUser
from app_management.models import (
    Language,
    ProblemCategory,
    Problem,
    Contest,
    ContestStatus,
    ContestProblem,
    SolutionReleasePolicy,
)
from app_oj.models import Submission, ContestRank, JudgeStatus

from django.core import serializers
from django.http import HttpResponse
import pytz
from django.utils import timezone
import logging
from django.views.decorators.csrf import csrf_exempt

from app_account.models import User

import openpyxl
from django.http import HttpResponse
import math
from openpyxl import load_workbook
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.contrib.sessions.models import Session
from .models import MySysOptions
from rest_framework.views import APIView
from rest_framework.response import Response
import json
from django.db.models import ManyToManyField
from django.db.models.fields.reverse_related import ForeignObjectRel

from django.http import QueryDict
from urllib.parse import urlencode

import pytz

# from pytz import timezone
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models import Q

import markdown
from app_oj.views import parse_problem_template
from app_oj.views import judge, judger_url

import requests
import base64

curr_timezone = timezone.get_current_timezone()

logger = logging.getLogger(__name__)
# logger.debug('debug')
# logger.info('info')
# logger.warning('warning')
# logger.error('error')
# logger.critical('critical')
# logger.exception('exception')


@staff_member_required
def management(request):

    return render(request, "app_management/home_management.html")


@staff_member_required
def problem_list(request):

    # if request.method == "GET":
    # 只使用GET模式即可
    # 獲取所有問題並按id降序排序
    problems = Problem.objects.all().order_by("-id")

    # 獲取多個類別過濾條件
    # Get selected categories from the request
    selected_categories = request.GET.getlist("selected_categories")

    # logger.info(selected_categories)

    if selected_categories:
        # 使用 Q 對象進行多個條件的 OR 過濾。使用 distinct() 方法來確保每個 problem 只出現一次
        query = Q()
        for category in selected_categories:
            query |= Q(categories__name=category)
        problems = problems.filter(query).distinct()

    # 獲取當前頁數，默認為第1頁
    page = request.GET.get("page", 1)

    # 使用Paginator進行分頁，每頁顯示15個問題
    paginator = Paginator(problems, 12)
    try:
        paginated_problems = paginator.page(page)
    except PageNotAnInteger:
        paginated_problems = paginator.page(1)
    except EmptyPage:
        paginated_problems = paginator.page(paginator.num_pages)

    # 章節主題清楚分類，可用空白隔開多個，例如:數學運算 判斷 迴圈 亂數 靜態方法 一維陣列 二維陣列 自訂類別 例外 檔案 繼承 抽象類別 介面 資料結構

    problem_categories = ProblemCategory.objects.all().order_by("id_seq")

    context = {
        "categories": problem_categories,
        "problems": paginated_problems,
        "selected_categories": selected_categories,
    }
    # 將分頁後的問題列表傳遞給模板
    return render(request, "app_management/problem_list.html", context)


@staff_member_required
def problem_create(request):

    if request.method == "GET":
        problem_json = {
            "template": """
//PREPEND BEGIN
//PREPEND END
//TEMPLATE BEGIN
//TEMPLATE END
//APPEND BEGIN
//APPEND END"""
        }
        problem_categories = ProblemCategory.objects.all().order_by("id_seq")
        languages = Language.objects.all()

        content = {
            "problem": problem_json,
            "categories": problem_categories,
            "languages": languages,
        }
        return render(request, "app_management/problem_create.html", content)

    elif request.method == "POST":
        problem_json = {
            "title": request.POST.get("title"),
            "description": request.POST.get("description"),
            "input_output_description": request.POST.get("input_output_description"),
            "std_input": request.POST.get("std_input").strip(),
            "std_output": request.POST.get(
                "std_output"
            ),  # 不能strip()有些題目的答案需要前面有空白
            "template": request.POST.get("template"),
            "sample_code": request.POST.get("sample_code"),
        }

        category_ids = request.POST.getlist("categories")
        # categories = ProblemCategory.objects.filter(id__in=category_ids)
        # problem_json["category"] = ' '.join(category.name for category in categories)

        language_id = request.POST.get("language")

        try:
            # 獲取 Problem 模型的所有欄位名稱
            fields = [field.name for field in Problem._meta.get_fields()]
            # logger.info(f"problem fields:{fields}")
            # logger.info(f"新增考題:{problem_json}")
            problem = Problem.objects.create(**problem_json)
            # logger.info(f"看到這行表示新增考題成功！")
            problem.categories.set(category_ids)
            problem.language = Language.objects.get(pk=language_id)
            problem.save()  # 再次保存 problem
            messages.success(request, "新增考題成功!")
            return redirect(reverse("problem_list"))
            # return redirect("problem_list")
        except Exception as e:
            logger.info(f"新增考題失敗: {e}")
            messages.error(request, "新增考題失敗!")
            return redirect("problem_create")  # Redirect back to create page on failure


@staff_member_required
def problem_duplicate(request, pk):
    problem = Problem.objects.get(pk=pk)
    original_categories = problem.categories.all()

    problem.title = "複製: " + problem.title
    problem.pk = (
        None  # This will create a new instance instead of updating the existing one
    )

    problem.save()  # Save the new instance to the database

    # When you duplicate the Problem instance, the ManyToManyField relationships are not automatically copied. You need to manually copy the categories field.
    problem.categories.set(original_categories)  # Copy the categories
    messages.success(request, "複製成功!")

    # Create a QueryDict to include selected categories and page number in the redirect URL
    prblm_list_filters = QueryDict(mutable=True)
    prblm_list_filters.update(request.GET)

    return redirect(f"{reverse('problem_list')}?{prblm_list_filters.urlencode()}")
    # return redirect('problem_list')


@staff_member_required
def contest_duplicate(request, pk):
    contest = get_object_or_404(Contest, pk=pk)

    # Create a new contest instance with the same attributes
    contest.pk = (
        None  # This will create a new instance instead of updating the existing one
    )
    contest.title = "複製: " + contest.title
    contest.save()  # Save the new instance to the database

    # Copy the ManyToManyField relationships by duplicating ContestProblem instances
    original_contest_problems = ContestProblem.objects.filter(contest=pk)
    for contest_problem in original_contest_problems:
        contest_problem.pk = None
        contest_problem.contest = contest
        contest_problem.save()

    messages.success(request, "複製成功!")

    # Create a QueryDict to include selected categories and page number in the redirect URL
    prblm_list_filters = QueryDict(mutable=True)
    prblm_list_filters.update(request.GET)

    return redirect(
        f"{reverse('contest_list_manage')}?{prblm_list_filters.urlencode()}"
    )
    # return redirect('contest_list_manage')


@staff_member_required
def problem_update(request, pk):
    problem_categories = ProblemCategory.objects.all().order_by("id_seq")
    languages = Language.objects.all()
    problem = get_object_or_404(Problem, id=pk)

    if request.method == "POST":
        problem.title = request.POST.get("title")
        problem.description = request.POST.get("description")
        problem.input_output_description = request.POST.get("input_output_description")
        problem.std_input = request.POST.get("std_input", "").strip()
        problem.std_output = request.POST.get(
            "std_output"
        ).rstrip()  # 不能strip()因為有些題目的答案需要前面有空白
        problem.template = request.POST.get("template")
        problem.sample_code = request.POST.get("sample_code")
        problem.note = request.POST.get("note")

        category_ids = request.POST.getlist("categories")

        language_id = request.POST.get("language")
        problem.language = Language.objects.get(pk=language_id)

        try:
            problem.save()
            problem.categories.set(
                category_ids
            )  # Update categories after saving the problem
            messages.success(request, "更新成功!")
        except Exception as e:
            logger.error(f"Fail to save: {e}")
            messages.error(request, "更新失敗!")

    category_ids = list(problem.categories.values_list("id", flat=True))
    content = {
        "problem": problem,
        "categories": problem_categories,
        "category_ids": category_ids,
        "languages": languages,
    }
    # 停留在更新頁面，沒有返回problem_list.html頁面
    return render(request, "app_management/problem_update.html", content)


@staff_member_required
def problem_delete(request, pk):

    # GET 模式
    problem = Problem.objects.get(id=pk)

    try:
        problem.delete()
    except:
        messages.success(request, "刪除失敗!")
    else:
        messages.success(request, "刪除成功!")

    # 回到problem_list頁面的準備動作:原來的頁次與題目類別參數。
    # Create a QueryDict to include selected categories and page number in the redirect URL
    # request.GET: <QueryDict: {'prblm_page_num': ['4'], 'selected_categories': ['靜態方法', '一維陣列']}>
    # prblm_list_filters: <QueryDict: {'prblm_page_num': ['4'], 'selected_categories': ['靜態方法', '一維陣列']}>

    # logger.info(f"request.GET: {request.GET}")
    # 方式1 一次取得
    prblm_list_filters = QueryDict(mutable=True)
    prblm_list_filters.update(request.GET)

    # 方式2 逐一取得
    """
    selected_categories = request.GET.getlist("selected_categories")
    page_number = request.GET.get("page", 1)
    
    prblm_list_filters = QueryDict(mutable=True)
    prblm_list_filters.update({'page': page_number})
    for category in selected_categories:
       prblm_list_filters.update({'selected_categories': category})
    
    # prblm_list_filters.update({'selected_categories': request.GET.get('selected_categories')}) # 不可行只會取得最後一個分類值
    """

    return redirect(f"{reverse('problem_list')}?{prblm_list_filters.urlencode()}")


@staff_member_required
def problem_belongs_to(request, pk):

    if request.method == "GET":
        problem = Problem.objects.get(pk=pk)
        contests = Contest.objects.all().order_by("-display_seq")
        # contests = Contest.objects.all().order_by("-id")
        # contests分頁處理
        page = request.GET.get("page", 1)
        paginator = Paginator(contests, 10)
        try:
            contests = paginator.page(page)
        except PageNotAnInteger:
            contests = paginator.page(1)
        except EmptyPage:
            contests = paginator.page(paginator.num_pages)

        response = {
            "contests": contests,
            "problem": problem,
            # 這裡的prblm_page_num selected_categories是為了點選確定按鈕後可以回到原本的頁面，若不使用session傳遞，則必須維持以下兩項參數
            # 在前端的problem_belongs_to_contest.html中，會有一個隱藏的input欄位，帶入prblm_page_num與selected_categories參數，透過此方式傳遞參數
            # "prblm_page_num": prblm_page_num,
            # "selected_categories": selected_categories,
        }

        #### 使用session傳遞參數，可以來回攜帶參數
        # Store the page number and selected categories in the session
        # 透過session傳遞參數原來problems列表的頁次還有題目類別過濾參數，帶到problem_belongs_to_contest.html挑選競賽，挑選完畢之後要將原來的頁碼與題目類別參數帶回到problem_list.html

        # 只在第一次進入時設定 session，之後的 GET 不再重複設定
        # Store initial parameters in session when first arriving from problem_list
        # Get the page number from the request 原來的problems paginator頁碼，名稱不要用page，因為會跟contests paginator的page衝突
        if "prblm_page_num" in request.GET:
            # 只在第一次進入時設定 session，之後的 GET 不再重複設定
            # Get the page number from the request 原來的problems paginator頁碼，名稱不要用page，因為會跟contests paginator的page衝突
            prblm_page_num = request.GET.get("prblm_page_num")
            selected_categories = request.GET.getlist("selected_categories")
            request.session["prblm_page_num"] = prblm_page_num
            request.session["selected_categories"] = selected_categories
        else:
            pass  # 選取競賽頁次，則不設定session參數(沒有prblm_page_num參數)，原來的session參數仍然存在
        # 渲染
        return render(
            request, "app_management/problem_belongs_to_contest.html", response
        )

    elif request.method == "POST":
        # For POST
        # logger.info("POST")
        # belong_to_which={}
        selected_contests = request.POST.getlist("belong_to_contests")
        # logger.info(f"belong_to_which: {belong_to_whichs}")
        problem = Problem.objects.get(pk=pk)

        # 選競賽可選多個，此處無法刪除隸屬競賽
        for selected_contest in selected_contests:
            contest = Contest.objects.get(pk=selected_contest)
            try:
                # 嘗試取得ContestProblem實例 若已經存在 就不做任何動作
                contest_problem = ContestProblem.objects.get(
                    problem=problem, contest=contest
                )
            except ContestProblem.DoesNotExist:
                contest_problem = ContestProblem.objects.create(
                    problem=problem, contest=contest
                )
                # logger.info(f"考題隸屬比賽: {contest.title}成功!")

        # 回到problem_list頁面的準備動作:原來的頁次與題目類別參數。
        # 從session取得參數
        prblm_page_num = request.session.get("prblm_page_num", 1)
        selected_categories = request.session.get("selected_categories", [])

        # 攜帶參數使用普通字典
        params = {}
        if prblm_page_num:
            params["page"] = prblm_page_num

        # 處理多值參數（如selected_categories）
        if selected_categories:
            # 方法1a：使用列表（urlencode會自動處理多值）
            params["selected_categories"] = selected_categories

        # logger.info(f"params: {params}")
        # logger.info(f"params urlencode: {urlencode(params, doseq=True)}")
        return redirect(f"{reverse('problem_list')}?{urlencode(params, doseq=True)}")

@staff_member_required
def contest_list_manage(request):

    # 只有GET的方式
    contests = Contest.objects.all().order_by("-display_seq")

    # Filter by contest public status: public_contest_filter
    public_contest_filter = request.GET.get("public_contest_filter")
    # logger.info(f"public_contest_filter: {public_contest_filter}")
    if public_contest_filter == "true":
        contests = contests.filter(is_public=True)
    elif public_contest_filter == "false":
        contests = contests.filter(is_public=False)

    # Filter by contest_status_filter
    contest_status_filter = request.GET.get("contest_status_filter")
    # logger.info(f"contest_status_filter: {contest_status_filter}")
    # 全部競賽狀態: 回傳值為 "None" 空字串
    if contest_status_filter and contest_status_filter != "None":
        contests = [
            contest for contest in contests if contest.status == contest_status_filter
        ]

    page = request.GET.get("page", 1)

    paginator = Paginator(contests, 10)
    try:
        contests = paginator.page(page)
    except PageNotAnInteger:
        contests = paginator.page(1)
    except EmptyPage:
        contests = paginator.page(paginator.num_pages)

    response = {
        "contests": contests,
        "SolutionReleasePolicy": SolutionReleasePolicy,
        "ContestStatus": ContestStatus,
        "public_contest_filter": public_contest_filter,
        "contest_status_filter": contest_status_filter,
    }
    return render(request, "app_management/contest_list_manage.html", response)


@staff_member_required
def contest_problems_manage(request, contest_pk):
    # GET
    # 前端contest_list_manage.html的使用者點選"+題目"按鈕時，會以GET帶入contest_pk參數
    if request.method == "GET":
        contest = Contest.objects.get(pk=contest_pk)
        contest_problems = ContestProblem.objects.filter(contest=contest).order_by(
            "id_prblm_in_contest"
        )

        # 以下3種寫法皆可
        # problems = contest.problems.all()
        # problems = contest.get_problems()
        problems = Problem.objects.filter(contests__id__contains=contest.id).order_by(
            "-created_at"
        )
        context = {
            "contest": contest,
            "contest_problems": contest_problems,
            "problems": problems,
        }
        return render(request, "app_management/contest_problems_manage.html", context)

    # 前端contest_problems_manage.html的使用者點選"確定修改編號與移除隸屬"按鈕時，會以POST帶入contest_pk參數
    elif request.method == "POST":

        # 欲修改的輸入考題編號
        new_ids_for_problem_in_contest = request.POST.getlist(
            "new_ids_for_problem_in_contest"
        )

        # 欲修改的考題problem_id
        problem_ids_to_be_modified = request.POST.getlist(
            "problem_ids_to_be_modified"
        )
        # print(new_ids_for_problem_in_contest)
        # print(problem_ids_to_be_modified)

        # 欲移除的編號proble_id
        removed_from_contest = request.POST.getlist("removed_from_contest")
        # print(removed_from_contest)

        contest = Contest.objects.get(pk=contest_pk)

        # 修改編號
        for idx in range(len(problem_ids_to_be_modified)):

            problem = Problem.objects.get(pk=problem_ids_to_be_modified[idx])

            contest_problem = ContestProblem.objects.get(
                problem=problem, contest=contest
            )
            contest_problem.id_prblm_in_contest = new_ids_for_problem_in_contest[idx]
            contest_problem.save()

        # 移除隸屬
        for idx in range(len(removed_from_contest)):
            problem = Problem.objects.get(pk=removed_from_contest[idx])
            contest_problem = ContestProblem.objects.get(
                problem=problem, contest=contest
            )
            contest_problem.delete()

        return redirect("contest_problems_manage", contest_pk=contest.pk)


@staff_member_required
def contest_create(request):

    if request.method == "POST":

        contest_json = {}
        contest_json["title"] = request.POST.get("title")
        contest_json["display_seq"] = request.POST.get("display_seq")
        contest_json["description"] = request.POST.get("description")

        # contest_json["start_time"] = request.POST.get('start_time')
        # contest_json["end_time"] = request.POST.get('end_time')

        start_time = datetime.strptime(request.POST.get("start_time"), "%Y-%m-%dT%H:%M")
        contest_json["start_time"] = timezone.make_aware(start_time)

        end_time = datetime.strptime(request.POST.get("end_time"), "%Y-%m-%dT%H:%M")
        contest_json["end_time"] = timezone.make_aware(end_time)

        # 若有被勾選回傳值(寫在value= 的值) "is_public_checked"，若沒有勾選擇回傳 None
        contest_json["is_public"] = (
            True if request.POST.get("is_public") == "is_public_checked" else False
        )
        # 若有被勾選回傳值(寫在value= 的值) "is_weighted_ranking_checked"，若沒有勾選擇回傳 None
        contest_json["is_weighted_ranking"] = (
            True
            if request.POST.get("is_weighted_ranking") == "is_weighted_ranking_checked"
            else False
        )

        # print(contest_json)

        try:
            Contest.objects.create(**contest_json)
            return redirect("contest_list_manage")
        except Exception as exception:
            # print("Fail to save!")
            # print(exception)
            logger.error("Fail to save contest!")
            logger.error(exception)

    else:
        contest_json = {}

        current_time = datetime.now()
        # formatted_time = current_time.strftime('%Y/%m/%d %H:%M:%S')
        # YYYY-MM-DD HH:MM 遇到空白被截斷??
        contest_json["start_time"] = current_time
        contest_json["end_time"] = current_time
        # 編號順序
        year = current_time.year
        month = current_time.month
        week_number = current_time.isocalendar()[1]
        day = current_time.day
        contest_json["display_seq"] = f"--w-"
        # contest_json["display_seq"] = f"{year}-1-w00-"
        # contest_json["display_seq"] = f"{year}-{month:02d}-w{week_number:02d}-{day:02d}-"

    return render(
        request, "app_management/contest_create.html", {"contest": contest_json}
    )


# 修改競賽細節 GET與POST兩種方式
@staff_member_required
def contest_update(request, pk):
    if request.method == "GET":
        # GET
        # 前端contest_list_manage.html的使用者點選"修改"按鈕時，會以GET帶入pk參數
        contest = Contest.objects.get(id=pk)
        # 回到原本的頁面需要的參數
        # 透過session傳遞參數 讓競賽頁次帶有這些參數，可以來回攜帶參數
        # ?page={{ contests.number }}{% if public_contest_filter %}&public_contest_filter={{ public_contest_filter }}{% endif %}{% if contest_status_filter %}&contest_status_filter={{ contest_status_filter }}{% endif %}
        # Store the page number and selected categories in the session
        request.session["public_contest_filter"] = request.GET.get(
            "public_contest_filter"
        )
        request.session["contest_status_filter"] = request.GET.get(
            "contest_status_filter"
        )
        request.session["page"] = request.GET.get("page", 1)

        return render(
            request, "app_management/contest_update.html", {"contest": contest}
        )

    # print("Update")
    if request.method == "POST":
        # POST
        # 前端contest_update.html的使用者點選"儲存"按鈕時，會帶入競賽細節修改的所有參數
        # print("POST")
        contest_json = {}
        contest_json["title"] = request.POST.get("title")
        contest_json["display_seq"] = request.POST.get("display_seq")
        contest_json["description"] = request.POST.get("description")

        #
        start_time = datetime.strptime(request.POST.get("start_time"), "%Y-%m-%dT%H:%M")
        contest_json["start_time"] = timezone.make_aware(start_time)

        end_time = datetime.strptime(request.POST.get("end_time"), "%Y-%m-%dT%H:%M")
        contest_json["end_time"] = timezone.make_aware(end_time)

        # 若有被勾選回傳值(寫在value= 的值) "is_public_checked"，若沒有勾選擇回傳 None
        contest_json["is_public"] = (
            True if request.POST.get("is_public") == "is_public_checked" else False
        )
        # 若有被勾選回傳值(寫在value= 的值) "is_weighted_ranking_checked"，若沒有勾選擇回傳 None
        contest_json["is_weighted_ranking"] = (
            True
            if request.POST.get("is_weighted_ranking") == "is_weighted_ranking_checked"
            else False
        )

        # print(contest_json)

        # 更新 Contest 物件
        Contest.objects.filter(pk=pk).update(**contest_json)

        # 網頁需要用到的參數 使用QueryDict來攜帶參數
        # 前端contest_update.html有重新取得與定義之前網頁過濾條件的參數，使用者點選"儲存"按鈕時，會再帶入進來
        # 使用 redirect() 函數時，伺服器會回應客戶端一個 HTTP 302 狀態碼和一個新的 URL。
        # page = request.POST.get("page")
        # public_contest_filter = request.POST.get("public_contest_filter")
        # contest_status_filter = request.POST.get("contest_status_filter")
        # 從 session 讀取參數
        page = request.session.get("page", 1)
        public_contest_filter = request.session.get("public_contest_filter")
        contest_status_filter = request.session.get("contest_status_filter")

        # 回到原本的頁面需要的參數
        contest_list_filters = QueryDict(mutable=True)
        # contest_list_filters.update(request.POST)
        contest_list_filters["page"] = page
        contest_list_filters["public_contest_filter"] = public_contest_filter
        contest_list_filters["contest_status_filter"] = contest_status_filter

        redirect_url = (
            f"{reverse('contest_list_manage')}?{contest_list_filters.urlencode()}"
        )
        # 更新成功後，回到競賽列表頁面  需要將之前的參數帶回去
        return redirect(redirect_url)

        # return HttpResponseRedirect(reverse('contest_list_manage'))
        # return redirect('contest_list_manage')
        # return render(request, "app_management/contest_list_manage.html")


@staff_member_required
def contest_owns_problems(request, contest_pk):

    # 只有用到GET
    if request.method == "GET":
        contest = Contest.objects.get(id=contest_pk)
        problems = Problem.objects.all().order_by("-id")

        contest_problems = ContestProblem.objects.filter(contest=contest)

        page = request.GET.get("page", 1)

        paginator = Paginator(problems, 10)
        try:
            problems = paginator.page(page)
        except PageNotAnInteger:
            problems = paginator.page(1)
        except EmptyPage:
            problems = paginator.page(paginator.num_pages)

        context = {
            "contest": contest,
            "problems": problems,
            "contest_problems": contest_problems,
        }
        return render(request, "app_management/contest_owns_problems.html", context)

    # 此部分使用者互動介面不好，被update_contest_owns_problems的ajax方式取代
    elif request.method == "POST":
        contest_owns = request.POST.getlist("contest_owns")
        contest = Contest.objects.get(pk=contest_pk)
        count = ContestProblem.objects.filter(contest=contest).count()
        # print("筆數:",count)

        for idx, problem_id in enumerate(contest_owns):
            problem = Problem.objects.get(pk=problem_id)

            try:
                contest_problem = ContestProblem.objects.get(
                    problem=problem, contest=contest
                )
                contest_problem.id_prblm_in_contest = "q"

                contest_problem.save()
            except ContestProblem.DoesNotExist:
                contest_problem = ContestProblem.objects.create(
                    problem=problem,
                    contest=contest,
                    id_prblm_in_contest="q" + str(count + idx + 1),
                )
                # print("新增:",contest_problem)

        # contestProblem, sucess = ContestProblem.objects.update_or_create(problem=problem, contest=contest, id_prblm_in_contest="q")
        return redirect("contest_problems_manage", contest_pk=contest.pk)


# ajax方式選取考題
@staff_member_required
def update_contest_owns_problems(request):
    contest_id = request.POST.get("contest_id")
    problem_id = request.POST.get("problem_id")
    contest = Contest.objects.get(pk=contest_id)
    problem = Problem.objects.get(pk=problem_id)
    is_selected = request.POST.get("is_selected")

    # logger.info(is_selected)
    # logger.info(type(is_selected))

    if is_selected == "True":
        contest_problem = ContestProblem.objects.create(
            problem=problem, contest=contest
        )
        logger.info("新增隸屬考題")
    else:
        contest_problem = ContestProblem.objects.filter(
            problem=problem, contest=contest
        )
        # contest_problem = ContestProblem.objects.get(problem=problem, contest=contest)
        contest_problem.delete()
        logger.info("刪除隸屬考題")

    logger.info(problem.contests.all().values("title"))

    # problem_contests = model_to_dict(problem.contests.all().values('title'))
    problem_contests = list(problem.contests.all().values("title"))
    # problem_contests = json.dumps( list(problem.contests.all().values('title')), ensure_ascii=False)

    #
    # model_to_dict(instance, fields=[], exclude=[])

    return JsonResponse(
        {
            "message": "Sucessfully update contest owns problems",
            "problem_contests": problem_contests,
        }
    )


@staff_member_required
def contest_delete(request, pk):
    contest = get_object_or_404(Contest, id=pk)
    prblm_list_filters = QueryDict(mutable=True)
    prblm_list_filters.update(request.GET)

    try:
        contest.delete()
        messages.success(request, "刪除比賽成功！")
    except Exception as e:
        logger.error(f"Error deleting contest {pk}: {e}")
        messages.error(request, "刪除比賽失敗！")

    redirect_url = f"{reverse('contest_list_manage')}?{prblm_list_filters.urlencode()}"
    return redirect(redirect_url)


@staff_member_required
def get_abnormal_users(request):

    # if request.is_ajax():

    list_ip_rejected_users = []
    for row in IPRegectedUser.objects.all():
        ip_rejected_users = {}
        ip_rejected_users["username"] = row.user.username
        ip_rejected_users["full_name"] = row.user.full_name
        ip_rejected_users["ip"] = row.ip
        ip_rejected_users["session_key"] = row.session_key
        ip_rejected_users["created_at"] = row.created_at.astimezone(
            curr_timezone
        ).strftime("%Y/%m/%d %H:%M:%S")

        list_ip_rejected_users.append(ip_rejected_users)

    list_session_warning_users = []
    for row in SessionWarningUser.objects.all():
        session_warning_users = {}
        session_warning_users["username"] = row.user.username
        session_warning_users["full_name"] = row.user.full_name
        session_warning_users["ip"] = row.ip
        session_warning_users["session_key"] = row.session_key
        session_warning_users["created_at"] = row.created_at.astimezone(
            curr_timezone
        ).strftime("%Y/%m/%d %H:%M:%S")
        list_session_warning_users.append(session_warning_users)

    content = {
        "ip_rejected_users": list_ip_rejected_users,
        "session_warning_users": list_session_warning_users,
    }

    return JsonResponse(content)


@staff_member_required
def delete_abnormal_users(request):
    IPRegectedUser.objects.all().delete()
    SessionWarningUser.objects.all().delete()
    return JsonResponse({"message": "Sucessfully delete abnormal users' records"})


@staff_member_required
def delete_user_sessions(request):
    """
    刪除所有非管理員的 sessions 記錄
    """
    # Session.objects.all().delete()

    # Get all sessions
    sessions = Session.objects.all()

    for session in sessions:
        # Decode the session to get the user ID
        session_data = session.get_decoded()
        user_id = session_data.get("_auth_user_id")

        if user_id:
            try:
                # Get the user associated with this session
                user = User.objects.get(id=user_id)
                # Check if the user is not a staff member
                if not user.is_staff:
                    # If not staff, delete the session
                    session.delete()
            except User.DoesNotExist:
                # In case the user does not exist, delete the session
                session.delete()

    logger.info("所有使用者的 sessions 已成功清除!")
    # messages.success(request, "所有使用者的 sessions 已成功清除!")
    return JsonResponse({"message": "所有使用者的 sessions 已成功清除。"})


@staff_member_required
def manage_contest(request):
    return render(request, "app_management/manage_oj_settings.html")


@staff_member_required
def ranking_download(request, contest_pk):

    contest = Contest.objects.get(pk=contest_pk)

    # 如果競賽還沒開始就不可下載
    if contest.status is ContestStatus.CONTEST_NOT_START:
        return redirect("contest_list_manage")

    staff = User.objects.filter(is_staff=True)

    # 排名資料表:按照答對題數降冪排列，若答對題數相同則按照總時間升冪排列
    contest_ranks = (
        ContestRank.objects.filter(contest=contest)
        .select_related("submitted_by")
        .exclude(submitted_by__in=staff)
        .order_by("-accepted_count", "total_time")
    )
    #
    # <QuerySet [(4, 2, 99, 4, 1, 591552, {'968': {'is_ac': True, 'ac_time': 591252.166756, 'error_count': 1, 'qz_prblm_id': 'qz-1'}, '963': {'is_ac': False, 'ac_time': 0, 'error_count': 1, 'qz_prblm_id': 'qz-1'}, '967': {'is_ac': False, 'ac_time': 0, 'error_count': 1, 'qz_prblm_id': 'qz-1'}})]>
    # contest_ranks = ContestRank.objects.filter(contest=contest).select_related("submitted_by").exclude(submitted_by__in=staff).order_by("-accepted_count", "total_time").values_list()

    contest_problems = ContestProblem.objects.filter(contest=contest).order_by(
        "id_prblm_in_contest"
    )

    numberOfproblems = contest_problems.count()
    # print(contest_problems)
    # print(contest_ranks)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="contest_ranking_data.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Ranking Data"

    # 根據 contest.is_weighted_ranking 決定是否包含加權總分欄位
    header = [
        "排名",
        "帳號",
        "姓名",
        "班級",
        "總分",
    ]

    # 如果啟用加權排名，將加權總分欄位添加到 header
    if contest.is_weighted_ranking:
        header.append("加權總分")

    # 添加其他欄位
    header.extend(
        [
            "答對總題數",
            "提交錯誤次數",
            "完成總時間(秒)",
            "平均完成時間(分鐘)",
        ]
    )

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # 題目題號: for是否有提交
    header_len = len(header)
    for idx, pb in enumerate(contest_problems, 1):
        cell = worksheet.cell(row=1, column=header_len + idx)
        cell.value = f"題號{pb.id_prblm_in_contest}"

    # 題目題號: for程式碼
    for idx, pb in enumerate(contest_problems, 1):
        cell = worksheet.cell(row=1, column=header_len + numberOfproblems + idx)
        cell.value = f"題號{pb.id_prblm_in_contest}"

    # 計算滿分有多少人和加權排名相關變數
    num_perfect_score = 0
    range_for_weighted_score = 0
    base_for_weighted_score = 0

    # 如果啟用加權排名，計算並保存加權總分
    if contest.is_weighted_ranking:
        # 計算滿分有多少人
        num_perfect_score = 0
        for record in contest_ranks:
            if record.accepted_count == numberOfproblems:
                num_perfect_score += 1
        # 計算加權總分用到的變數
        range_for_weighted_score = math.ceil(
            100 / numberOfproblems / 2
        )  # 5題，每個題目20分，50%的範圍: 100/5/2=10
        base_for_weighted_score = math.ceil(100 - range_for_weighted_score)  # 90分

    # 用於跟踪列數的偏移量 - 根據加權排名是否啟用而變化
    weighted_offset = 1 if contest.is_weighted_ranking else 0

    for row_idx, record in enumerate(contest_ranks, 1):
        # print(record.submitted_by.username)
        # print(record.submitted_by.full_name)
        rank = row_idx
        worksheet.cell(row=1 + row_idx, column=1).value = rank  # 排名
        worksheet.cell(row=1 + row_idx, column=2).value = record.submitted_by.username
        worksheet.cell(row=1 + row_idx, column=3).value = record.submitted_by.full_name
        worksheet.cell(row=1 + row_idx, column=4).value = record.submitted_by.user_class

        # 計算並保存總分
        # print(numberOfproblems)
        # print(record.accepted_count)
        total = math.ceil(100 * (record.accepted_count / numberOfproblems))
        worksheet.cell(row=1 + row_idx, column=5).value = total

        # 計算並保存加權總分
        if contest.is_weighted_ranking:
            if record.accepted_count == numberOfproblems and num_perfect_score > 0:
                # 只有當有滿分者時才需計算加權分數公式
                weighted_total = base_for_weighted_score + round(
                    range_for_weighted_score
                    * ((num_perfect_score - rank + 1) / num_perfect_score)
                )
            else:
                # 不是滿分或沒有人滿分時，就用原始分數
                weighted_total = total

            worksheet.cell(row=1 + row_idx, column=6).value = weighted_total

        # 保存答題情況 - 注意列數偏移
        worksheet.cell(row=1 + row_idx, column=5 + weighted_offset + 1).value = (
            record.accepted_count
        )
        worksheet.cell(row=1 + row_idx, column=5 + weighted_offset + 2).value = (
            record.submission_count - record.accepted_count
        )
        worksheet.cell(row=1 + row_idx, column=5 + weighted_offset + 3).value = (
            record.total_time
        )

        # 計算平均時間
        if record.total_time != 0:
            avg_time = round(record.total_time / 60 / record.accepted_count)
        else:
            avg_time = 0
        worksheet.cell(row=1 + row_idx, column=5 + weighted_offset + 4).value = avg_time

        # 每個題目是否有提交(1表示有提交，0表示沒有提交)
        submission_info = record.submission_info
        for col_idx, pb in enumerate(contest_problems, 1):
            # print("考題編號",pb.problem.pk)
            is_accepted = 0
            for key in submission_info:
                # print(key)
                # print(type(key))
                # print(type(pb.problem.pk))
                if str(pb.problem.pk) == key:
                    # print(submission_info[key]['is_ac'])
                    if submission_info[key]["is_ac"] == True:
                        is_accepted = 1
            # print(pb.problem.pk, pb.id_prblm_in_contest, is_accepted)
            worksheet.cell(row=1 + row_idx, column=header_len + col_idx).value = (
                is_accepted
            )

        # 保存提交程式碼
        # user = User.objects.get(id=record.submitted_by.id)
        user = record.submitted_by
        # logger.info(user.username)
        for col_idx, pb in enumerate(contest_problems, 1):
            problem = Problem.objects.get(pk=pb.problem.pk)
            # logger.info(problem.title)
            # submission = Submission.objects.get(problem=problem, contest=contest, submitted_by=user)
            # logger.info(submission)
            # logger.info(submission.source_code)
            # worksheet.cell(row=1+row_idx, column=header_len+numberOfproblems+1+col_idx).value = submission.source_code
            try:
                submission = Submission.objects.get(
                    problem=problem, contest=contest, submitted_by=user
                )
                # logger.info(submission.source_code)
                worksheet.cell(
                    row=1 + row_idx, column=header_len + numberOfproblems + col_idx
                ).value = submission.source_code
            except Submission.DoesNotExist:
                worksheet.cell(
                    row=1 + row_idx, column=header_len + numberOfproblems + col_idx
                ).value = ""

    workbook.save(response)

    # messages.success(request, "下載成功!")
    return response


"""
當你使用 Contest.objects.get(id=contest_id) 來獲取 Contest 實例時，如果你將一個沒有時區信息的 datetime 賦值給它的 start_time 屬性並保存，Django 會發出警告，因為時區支持是啟用的。

然而，當你使用 Contest.objects.filter(pk=pk).update(**contest_json) 時，更新是直接在數據庫中進行的，沒有將模型實例加載到內存中。這樣就繞過了 Django 的時區檢查，直接更新數據庫字段，因此你不會看到警告。
"""


@staff_member_required
def update_contest_start_time(request):
    contest_id = request.POST.get("contest_id")
    start_time = request.POST.get("start_time")

    contest = Contest.objects.get(id=contest_id)

    # 解析start_time並添加時區信息
    # Parse start_time and add timezone information
    start_time = datetime.strptime(start_time, "%Y-%m-%dT%H:%M")
    start_time = timezone.make_aware(start_time)

    contest.start_time = start_time

    contest.save()
    contest = Contest.objects.get(id=contest_id)

    # contest_status = contest.status_label() # NOT_STARTED, ENDED, UNDERWAY
    response = {
        "response": "update sucess",
        "contest_status": contest.status,
        "contest_status_label": contest.status_label(),
    }

    return JsonResponse(response)
    # return redirect('contest_list_manage')


@staff_member_required
def update_contest_end_time(request):
    contest_id = request.POST.get("contest_id")
    end_time = request.POST.get("end_time")

    contest = Contest.objects.get(id=contest_id)
    # 解析end_time並添加時區信息
    end_time = datetime.strptime(end_time, "%Y-%m-%dT%H:%M")
    contest.end_time = timezone.make_aware(end_time)
    contest.save()

    # logger.info(f"contest.end_time:{contest.end_time}")
    contest = Contest.objects.get(id=contest_id)

    # contest_status = contest.status  # NOT_STARTED, ENDED, UNDERWAY
    response = {
        "response": "update sucess",
        "contest_status": contest.status,
        "contest_status_label": contest.status_label(),
    }
    return JsonResponse(response)


@staff_member_required
def update_contest_publicity(request):
    contest_id = request.POST.get("contest_id")
    is_public = request.POST.get("is_public")

    contest = Contest.objects.get(id=contest_id)
    contest.is_public = is_public
    contest.save()
    return JsonResponse({"response": "update sucess"})


@staff_member_required
@csrf_exempt
def update_solution_release_policy(request):
    if request.method == "POST":
        data = json.loads(request.body)
        contest_id = data.get("contest_id")
        solution_release_policy = data.get("solution_release_policy")

        try:
            contest = Contest.objects.get(id=contest_id)
            contest.solution_release_policy = solution_release_policy
            # 這樣也OK
            # contest.solution_release_policy = SolutionReleasePolicy(solution_release_policy)
            # logger.info(solution_release_policy)
            # logger.info(SolutionReleasePolicy(solution_release_policy))
            # logger.info(contest_id)
            # logger.info(contest.solution_release_policy)
            contest.save()
            return JsonResponse({"success": True})
        except Contest.DoesNotExist:
            return JsonResponse({"error": False, "error": "Contest not found"})
        except Exception as e:
            return JsonResponse({"error": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


# pip install djangorestframework==3.11.1
class WebsiteConfigAPI(APIView):

    def get(self, request):
        sys_options = {
            item.option_name: item.option_value for item in MySysOptions.objects.all()
        }
        # sys_options = MySysOptions.objects.values()
        return Response(sys_options)

    def post(self, request):

        option_names = [
            "forbid_login_session",
            "forbid_login_ip",
            "allowed_ip_ranges",
            "forbid_password_change",
            "allow_user_register",
            "hide_submitted_code",
        ]

        for option_name in option_names:
            op = MySysOptions.objects.get(option_name=option_name)
            option_value = request.POST.get(option_name)
            if option_name == "allowed_ip_ranges":
                # ipranges = [{'ip':ip} for ip in request.POST.get('ipranges').split(";")]
                option_value = [ip.strip() for ip in option_value.split(",") if ip]
            op.option_value = option_value
            op.save()
        return Response("success")


@staff_member_required
def import_users_from_excel(request):
    if request.method == "POST":
        fields = ["username", "full_name", "password", "user_class"]
        data = {}

        try:
            excel_file = request.FILES["excel_file"]
            wb = load_workbook(excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=False):
                # logger.info(row)
                for idx, cell in enumerate(row):
                    data[fields[idx]] = cell.value

                try:
                    user = User.objects.get(username=data["username"])
                    user.password = make_password(data["password"])
                    # user.set_password(data['password'])
                    user.full_name = data["full_name"]
                    user.user_class = data["user_class"]
                    user.save()
                except User.DoesNotExist:
                    user = User.objects.create_user(**data)
                    # user = User.objects.create_user(username=data['username'], password=data['password'])
                    # user.full_name = data['full_name']
                    # user.user_type = data['user_type']
                    user.save()
            messages.success(request, "匯入使用者成功!")
        except Exception as e:
            messages.error(
                request, "錯誤:Excel的欄位必須正確且內容必須為純文字,或是檔案讀取錯誤!"
            )

        return redirect("import_users_from_excel")

    return render(request, "app_management/import_users_from_excel.html")


@staff_member_required
def problem_submit(request, pk):

    problem = Problem.objects.get(pk=pk)

    # 此版本在後端渲染，比在前端渲染快速，但是須設定```區域的格式: fenced_code 與 顏色codehilite顯示
    problem.description = markdown.markdown(
        problem.description, extensions=["fenced_code", "codehilite"]
    )
    problem.input_output_description = markdown.markdown(
        problem.input_output_description, extensions=["fenced_code", "codehilite"]
    )

    return render(request, "app_management/problem_submit.html", {"problem": problem})


def stringToBase64(src):
    return base64.b64encode(src.encode("utf-8")).decode("ascii")


def base64ToString(src):
    return base64.b64decode(src).decode("utf-8")


# 前端提交程式碼給Judger
@staff_member_required
@csrf_exempt
def submit_to_judger(request):
    problem_id = request.POST.get("problem_id")
    problem = Problem.objects.get(id=problem_id)

    # stdin必須要轉換編碼否則中文輸入Judger會報錯
    # 這裡主要檢查 stdin is None 轉換編碼會噴錯
    # 不必檢查了-->匯入題目時已經解決了，不會有此情況產生。

    """
    if stdin=='None':
        stdin=""
        logger.info("stdin是None字串")
    elif len(stdin.strip())==0:
        stdin=""
        logger.info("stdin是空字串")
    # 若為None type要轉換編碼會報錯，因此將其置換為空字串
    elif stdin is None:
        stdin=""
        logger.info("stdin是None型態，轉換為空字串")
    """

    # 一定要有base64編碼，否則有中文字會噴錯!
    # expectedOutput = stringToBase64(problem.std_output) #'5' #problem.expected_output
    # source_code = request.POST.get("source_code")
    # stdin = problem.std_input #'2 3' #problem.stdin
    # expectedOutput = problem.std_output #'5' #problem.expected_output

    # if (len(std_inputs) != len(std_outputs)):
    #     logger.error("測試案例數量不一致!")
    #     return JsonResponse({"message": "測試案例數量不一致!"})

    # test_case_list = []
    # for idx in range(len(std_inputs)):
    #     std_input = std_inputs[idx].strip() # 可以移除每組前後的空白和換行符
    #     std_output = std_outputs[idx].lstrip('\r\n') # 這裡要移除切割測資之後，第二個測資的最左側多一個換行符\r\n
    #     logger.info(f"std_input:{std_input}")
    #     logger.info(f"std_output:{std_output}")
    #     std_input = stringToBase64(std_input)
    #     std_output = stringToBase64(std_output)

    #     test_case_list.append({
    #         "stdin": std_input,  # 每組資料作為一個 stdin
    #         "expected_output": std_output  # 每組對應的預期輸出
    #     })
    # logger.info(test_case_list)
    # v1.14版才會有多筆測試案例，請期待!
    # test_case_list =[{
    #     "stdin": stdin,
    #     "expected_output":expectedOutput
    # }]
    # test_case_list =[{
    #     "stdin": problem.std_input,
    #     "expected_output":problem.std_output
    # }]

    source_code = request.POST.get("source_code")

    if problem.template.strip():
        template = parse_problem_template(problem.template)
        source_code = f"{template['prepend']}\n{source_code}\n{template['append']}"

    # 一定要base64編碼，否則有中文字會噴錯!
    source_code = stringToBase64(source_code)

    # 送給Judger執行
    judge_response_info = {}
    # 存放提交結果的 ID 列表
    submission_tokens = []
    judge_response_info["result_status"] = "SubmittingToJudgerOK"  # 處理中
    logger.info("準備好了，要丟給Judger!!!")

    #
    test_cases = problem.get_test_cases()

    # 提交每組測資
    for test_case in test_cases:
        # 取得提交結果的 token
        try:
            logger.info("丟給Judger!!!")
            logger.info(test_case["stdin"])
            task_token = judge(
                source_code,
                test_case["stdin"],
                test_case["expected_output"],
                problem.language.judge_id,
            )
            submission_tokens.append(str(task_token))
        except Exception as error:
            judge_response_info["result_status"] = "SubmittingToJudgerError"  #
            logger.error("丟給Judger，發生例外了!")
            logger.error(error)
            break

    # print("this is token:",task_token)
    judge_response_info["task_token"] = submission_tokens
    logger.info(judge_response_info)
    return JsonResponse(judge_response_info)


"""
{
    "source_code": "#include <stdio.h>\nint main() { int a, b; scanf(\"%d %d\", &a, &b); printf(\"%d\\n\", a + b); return 0; }",
    "language_id": 50,  // C語言的ID
    "stdin": "3 5",
    "expected_output": "8\n",
    "test_cases": [
        {
            "stdin": "3 5",
            "expected_output": "8\n"
        },
        {
            "stdin": "-1 4",
            "expected_output": "3\n"
        },
        {
            "stdin": "100 200",
            "expected_output": "300\n"
        }
    ]
}
"""


# 前端取得Judger的執行結果
@staff_member_required
@csrf_exempt
def get_manager_submission_result(request):
    submission_tokens = eval(request.POST.get("task_token"))
    logger.info(f"tokens: {submission_tokens}")

    # 存放提交結果的 ID 列表
    judge_response_info = {}
    for task_token in submission_tokens:
        try:
            # 取結果
            # logger.info("取得Judger結果!!!")
            # logger.info(task_token)
            judge_result = requests.get(
                f"{judger_url}/{task_token}?base64_encoded=true"
            ).json()
            # logger.info(judge_result)
        except Exception as error:
            logger.error("取得Judger結果，發生例外了!")
            logger.error(error)
            judge_response_info["result_status"] = "GetJudgerResultError"
            return JsonResponse(judge_response_info)

        # 取結果
        # result = requests.get(f'{judger_url}/{task_token}')
        # result = requests.get(f'{judger_url}/{task_token}?base64_encoded=true?header={"Content-Type": "application/json"}')
        # judge_result = requests.get(f'{judger_url}/{task_token}?base64_encoded=true').json()
        # result = requests.get(f'{judger_url}/{task_token}?base64_encoded=false')
        # print(response.json())
        # print(result.text)
        # logger.info(judge_result)

        status = judge_result["status"]["id"]

        # 若還沒有獲得執行結果，則返回"JudgerInQueueOrProcessing"
        if status == JudgeStatus.IN_QUEUE or status == JudgeStatus.PROCESSING:
            judge_response_info["result_status"] = "JudgerInQueueOrProcessing"
            return JsonResponse(judge_response_info)
        elif judge_result["status"]["id"] == JudgeStatus.ACCEPTED:
            # 若是AC，則繼續取下一個結果，直到全部取完
            continue
        elif judge_result["status"]["id"] == JudgeStatus.WRONG_ANSWER:
            # 若有錯誤，則返回錯誤訊息
            judge_response_info["result_status"] = "JudgerSUCCESS"
            # 取得錯誤代碼的描述
            judge_response_info["judge_status_description"] = judge_result["status"][
                "description"
            ]
            judge_response_info["judge_compile_output"] = (
                "未通過~~有部分或全部測資答案錯誤，請再仔細檢查程式碼!"
            )
            return JsonResponse(judge_response_info)
        else:
            # 若有錯誤，則返回錯誤訊息
            judge_response_info["result_status"] = "JudgerSUCCESS"
            # 取得錯誤代碼的描述
            judge_response_info["judge_status_description"] = judge_result["status"][
                "description"
            ]
            # 取得編譯錯誤訊息
            compile_output = judge_result["compile_output"]
            if compile_output is not None:
                compile_output = base64ToString(compile_output)
            judge_response_info["judge_compile_output"] = compile_output
            return JsonResponse(judge_response_info)

    # 所有都Acctepted 的情況
    # 錯誤訊息judge_result['status']['description']
    judge_response_info["result_status"] = "JudgerSUCCESS"
    judge_response_info["judge_status_description"] = judge_result["status"][
        "description"
    ]

    judge_response_info["judge_compile_output"] = "通過~~全部測資都正確!"

    return JsonResponse(judge_response_info)


@staff_member_required
def export_all_scores_to_excel(request):

    # 匯出
    httpResponse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    httpResponse["Content-Disposition"] = (
        'attachment; filename="exported_all_sudent_scores.xlsx"'
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Ranking Data"

    # 取得所有使用者的成績資料
    users = User.objects.filter(is_staff=False).order_by("user_class", "username")
    contests = Contest.objects.filter(is_public=True).order_by("display_seq")
    staff = User.objects.filter(is_staff=True)

    start_row = 2
    start_col = 4
    worksheet.cell(row=2, column=1, value="序號")
    worksheet.cell(row=2, column=2, value="學號")
    worksheet.cell(row=2, column=3, value="姓名")
    worksheet.cell(row=2, column=4, value="班別")

    # 計算每個競賽的欄位位置並寫入標題
    col_position = start_col

    for contest_num, contest in enumerate(contests, start=1):

        # 寫入競賽標題
        worksheet.cell(row=1, column=col_position + 1, value=contest.title)
        # 寫入總分標題
        worksheet.cell(row=2, column=col_position + 1, value="總分")

        # 如果啟用加權排名，才顯示加權總分欄位
        if contest.is_weighted_ranking:
            worksheet.cell(row=2, column=col_position + 2, value="加權總分")
            logger.info(f"{contest.title}--啟用加權排名，顯示加權總分欄位")

        # 排名資料表:按照答對題數降冪排列，若答對題數相同則按照總時間升冪排列
        all_users_contest_ranks = (
            ContestRank.objects.filter(contest=contest)
            .select_related("submitted_by")
            .exclude(submitted_by__in=staff)
            .order_by("-accepted_count", "total_time")
        )

        # 計算競賽的題目數量
        numberOfproblems = ContestProblem.objects.filter(contest=contest).count()

        # 計算滿分有多少人
        if contest.is_weighted_ranking:
            num_perfect_score = 0
            for record in all_users_contest_ranks:
                if record.accepted_count == numberOfproblems:
                    num_perfect_score += 1
            # logger.info(num_perfect_score)

        # 處理每個使用者的成績資料
        for user_num, user in enumerate(users, start=1):

            # 寫入學生基本資料
            worksheet.cell(row=start_row + user_num, column=1, value=user_num)
            worksheet.cell(row=start_row + user_num, column=2, value=user.username)
            worksheet.cell(row=start_row + user_num, column=3, value=user.full_name)
            worksheet.cell(row=start_row + user_num, column=4, value=user.user_class)

            # 取得使用者的成績資料
            user_constest_rank = ContestRank.objects.filter(
                contest=contest, submitted_by=user
            ).first()

            # logger.info(user_constest_rank)

            # 設定初始總分
            total = 0
            weighted_total = 0
            if user_constest_rank is not None:
                # 計算總分
                total = math.ceil(
                    100 * (user_constest_rank.accepted_count / numberOfproblems)
                )

                # 只在啟用加權排名時才計算加權總分
                if contest.is_weighted_ranking:
                    # 計算排名
                    user_rank = 0
                    for index, rank in enumerate(all_users_contest_ranks, start=1):
                        # logger.info(rank.submitted_by)
                        if rank.submitted_by == user:
                            user_rank = index
                            break
                    # logger.info(user_rank)

                    # 計算加權總分用到的變數 最多分數是100分，除以題目數量，除以2得到加權分數的範圍
                    # 例如5題，每個題目20分，50%的範圍: 100/5/2=10
                    range_for_weighted_score = math.ceil(100 / numberOfproblems / 2)
                    base_for_weighted_score = math.ceil(100 - range_for_weighted_score)
                    # 計算加權總分
                    if (
                        user_constest_rank.accepted_count == numberOfproblems
                        and num_perfect_score > 0
                    ):
                        weighted_total = base_for_weighted_score + round(
                            range_for_weighted_score
                            * ((num_perfect_score - user_rank + 1) / num_perfect_score)
                        )
                    else:
                        weighted_total = total

            # 寫入總分
            worksheet.cell(
                row=start_row + user_num,
                column=col_position + 1,
                value=total,
            )
            # 如果啟用加權排名，才寫入加權總分
            if contest.is_weighted_ranking:
                worksheet.cell(
                    row=start_row + user_num,
                    column=col_position + 2,
                    value=weighted_total,
                )

        # 計算下一個競賽的欄位位置
        # logger.info(f"競賽編號: {contest_num}, 競賽標題: {contest.title}, 欄位位置: {col_position}")
        contest_col_offset = 2 if contest.is_weighted_ranking else 1
        col_position += contest_col_offset
        # logger.info(f"下一個競賽的欄位位置: {col_position}")

    workbook.save(httpResponse)
    return httpResponse


@staff_member_required
def import_problems_from_excel(request):
    if request.method == "POST":
        # id	category	title	description	input_output_description	std_input	std_output	sample_code	created_at	modified_at	is_sample_code_visible	is_public	languages	template	note

        # Define the fields to be imported
        # fields = [field.name for field in Problem._meta.get_fields() if field.name != 'categories']

        fields = [
            field.name
            for field in Problem._meta.get_fields()
            if not isinstance(field, (ForeignObjectRel)) and field.name != "id"
        ]

        data = {}
        try:
            excel_file = request.FILES["excel_file"]
            wb = load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=False):
                for idx, cell in enumerate(row):
                    # 如果是空的Excel欄位值，將其設定為""空字串比較好，系統預設會設定為"None"字串，這樣很麻煩!
                    # std_input欄位有時候會是空字串，匯出Excel時會使得Excel欄位值為空的，因此匯入時需要特別處理。

                    # Get the field name from the header row
                    column_letter = cell.column_letter
                    field_name = ws[f"{column_letter}1"].value

                    logger.info(field_name)
                    if field_name == "categories":
                        if cell.value is not None:
                            category_names = [
                                name.strip() for name in cell.value.split(",")
                            ]
                    elif field_name == "language":
                        if cell.value is not None:
                            language_name = cell.value.strip()
                            language = Language.objects.get(name=language_name)
                            data[field_name] = language
                    else:
                        data[field_name] = cell.value if cell.value is not None else ""
                    logger.info(cell.value)

                # Remove the id field because it is auto-generated
                # id欄位不需要，因為是自動產生的
                data.pop("id", None)
                logger.info(data)
                # Create the Problem object
                problem = Problem.objects.create(**data)
                # problem.save()
                logger.info(problem)
                # Process categories
                if category_names:
                    categories = []
                    for name in category_names:
                        category, created = ProblemCategory.objects.get_or_create(
                            name=name
                        )
                        categories.append(category)
                    problem.categories.set(categories)

            messages.success(request, "匯入考題成功!")
        except Exception as e:
            messages.error(request, e)
            messages.error(request, "匯入錯誤:注意欄位必須正確且內容必須為純文字!")
    return render(request, "app_management/import_problems_from_excel.html")


@staff_member_required
def export_problems_to_excel(request):

    if request.method == "POST":

        # 欲匯出的編號
        # 前端會將有勾選的input value送到後端
        # <input class="form-check-input" type="checkbox" name="problem_exported" value="{{ problem.pk }}">
        exported_problems = request.POST.getlist("problem_exported")
        logger.info(exported_problems)

        # 匯出
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="oj_exported_problems.xlsx"'
        )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Ranking Data"

        # Write header row
        # fields = ['id','category', 'title', 'description', 'input_output_description','std_input','std_output','sample_code','created_at','modified_at','is_sample_code_visible','is_public','languages','template','note']

        # ['contests', 'contestproblem', 'submission', 'id', 'category', 'title', 'description', 'input_output_description', 'std_input', 'std_output', 'sample_code', 'created_at', 'modified_at', 'is_sample_code_visible', 'is_public', 'languages', 'template', 'note', 'categories']
        # skipped_fields = ['contestproblem']
        # 取得所有欄位名稱 但是要排除掉不需要的欄位 例如: contests contestproblem 外鍵欄位
        fields = [
            field.name
            for field in Problem._meta.get_fields()
            if not isinstance(field, (ForeignObjectRel))
        ]
        logger.info(fields)

        for col_num, column_title in enumerate(fields, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        for row_idx, problem_pk in enumerate(exported_problems, 1):
            problem = Problem.objects.get(pk=problem_pk)
            for col_idx, column_name in enumerate(fields, 1):
                logger.info(column_name)
                if column_name == "categories":
                    # Handle the many-to-many field separately
                    categories = problem.categories.all()
                    category_names = ", ".join(
                        [category.name for category in categories]
                    )
                    value = category_names
                    logger.info(value)
                elif column_name == "contests":
                    # Handle the many-to-many field separately
                    contests = problem.contests.all()
                    contest_names = ", ".join([contest.title for contest in contests])
                    value = contest_names
                    logger.info(value)
                else:
                    value = getattr(problem, column_name)

                if isinstance(value, datetime):
                    value = value.astimezone(taipei_timezone).strftime(
                        "%Y/%m/%d %H:%M:%S"
                    )
                if value == []:
                    value = ""

                logger.info(value)
                worksheet.cell(row=1 + row_idx, column=col_idx).value = value

        workbook.save(response)
        return response
        # return redirect('problem_list')

    else:
        problems = Problem.objects.all().order_by("-id")
        # problems = Problem.objects.all().order_by("-created_at")
        page = request.GET.get("page", 1)
        paginator = Paginator(problems, 25)
        try:
            problems = paginator.page(page)
        except PageNotAnInteger:
            problems = paginator.page(1)
        except EmptyPage:
            problems = paginator.page(paginator.num_pages)
        context = {"problems": problems}
        return render(request, "app_management/export_problems_to_excel.html", context)
